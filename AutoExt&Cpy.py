# Block 1: Impporting all the necessary libraries/modules for this project
import numpy as np
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
import os
import time
from datetime import datetime,timedelta


# Block 2: This project requires interactions with the CRMS website so we require webdriver and website variable assigned
website='https://crms.suzlon.com/'
path=r'E:\chromedriver\chromedriver'


# Block 3: Changing the settings of the chrome to specify the download directory on the the system
chrome_options=webdriver.ChromeOptions()
download_dir='D:\BHUPEN PCC\client data\CRMS Daily\savla\CRMS download daily'
chrome_options.add_experimental_option('prefs',{'download.default_directory':download_dir})
driver=webdriver.Chrome(options=chrome_options)
driver.get(website)


# Block 4: Entering the website by filling in the log details and clicking login button using selenium
# filling login page and entering the site
log_in=driver.find_element(By.ID,'txtUserId')
password=driver.find_element(By.ID,'txtPassword')
log_in.send_keys('*****')                                        # Info redacted for privacy reasons
password.send_keys('*********')
#clicking login button
login_button=driver.find_element(By.ID,'img_login')
login_button.click()


# Block 5: Navigating inside the website
driver.get('https://crms.suzlon.com/aspx/DailyGenerationReport.aspx')
Customer_dd=Select(driver.find_element(By.CSS_SELECTOR,'[name="ctl00$ContentPlaceHolder1$DDLCustomer"]'))
Customer_dd.select_by_visible_text('***** ******** ** ****')        # Info redacted for privacy reasons

# Using master worksheet to finf out latest date of the data retrieved then using that dsate to download further data from the site
from openpyxl import load_workbook,workbook
master_wb=openpyxl.load_workbook(r"D:\BHUPEN PCC\client data\CRMS Daily\Gee Cee CRMS\WTG Analysis - Copy.xlsm",data_only=True)
master_wb_ld=master_wb['LatestDate'] #master_wb_gen is a dataframe
latest_dates={}
for row in range(3,8):
    for col in range(3,5):
        cell=master_wb_ld.cell(row=row,column=col)
        if cell.value is not None:  
            latest_dates[master_wb_ld.cell(row=row,column=3).value]=master_wb_ld.cell(row=row,column=4).value
            break
latest_date_J224= latest_dates['J224'] + timedelta(days=1)
latest_date_J224_in=latest_date_J224.strftime("%A, %B %d, %Y")
print(latest_date_J224_in)                                        # For debugging purpose

# Filling 'from' box of the date section of webpage
# Date input from to extract data from CRMS
css_selector=f'[title="{(latest_date_J224_in)}"]'
month=f'//div[@title="{latest_date_J224.strftime("%B, %Y")}"]'
print(month)
print(css_selector)
calendar_button=driver.find_element(By.ID,"ContentPlaceHolder1_imgCalFromDate")
calendar_button.click()
calender1_title_button=driver.find_element(By.ID,"ContentPlaceHolder1_calendarButtonExtender_title")
calender1_title_button.click()
from_date_input_month=driver.find_element(By.XPATH,month)
from_date_input_month.click()
from_date_input=driver.find_element(By.CSS_SELECTOR,css_selector)
from_date_input.click()

# Filling 'to' box of the date section of webpage
calender2_button=driver.find_element(By.ID,"ContentPlaceHolder1_imgCalToDate")
calender2_button.click()
today_button=driver.find_element(By.ID,"ContentPlaceHolder1_CalendarExtender1_today")
today_button.click()

# The webpage does not generate any table on the webpage itself rather we have to download the excel file from the link and analyse after download
view_report_button=driver.find_element(By.ID,"ContentPlaceHolder1_BtnViewRpt")
view_report_button.click()
excel_dwnld_button=driver.find_element(By.ID,"ContentPlaceHolder1_gvDailyGenData_lnkbtnFile_0")
excel_dwnld_button.click()


# Block 6: Changing the file name client wise using os module of python
#1. changing latest downloaded file name
dir_ch_fname='D:\BHUPEN PCC\client data\CRMS Daily\savla\CRMS download daily'
dict_fname_change={}
date_list=[]
for file in os.scandir(dir_ch_fname):
  fm_date= time.strftime("%y/%m/%d",time.strptime(time.ctime(os.path.getmtime(file.path))))
  date_list.append(fm_date)
  dict_fname_change[fm_date]=file.path
latest_date=max(date_list)
print(dict_fname_change[latest_date])
  
  #2. Changing the file name
  today=datetime.now()
  yesterday=today-timedelta(days=1)
  old_name=dict_fname_change[latest_date]
  new_name="D:\BHUPEN PCC\client data\CRMS Daily\savla\CRMS download daily\savt"+yesterday.strftime("%d-%m-%y")+".xls"
  os.rename(old_name,new_name)


# Block 7: Updating master datasheet (SADLY THIS WORKS BUT NOT AS EXPECTED-->VBA is better for this scenario)
from openpyxl import load_workbook,workbook
master_wb=load_workbook(r"D:\BHUPEN PCC\client data\CRMS Daily\Gee Cee CRMS\WTG Analysis - Copy.xlsm")
daily_wb=load_workbook(r"D:\BHUPEN PCC\client data\CRMS Daily\savla\CRMS download daily\savt23-03-24.xls") 

daily_wb_gen=daily_wb['Generation Data'] #Generation sheet to be copied
master_wb_gen=master_wb['Generation'] #to be copied into this sheet of master wb

for row in daily_wb_gen.iter_rows(min_row=2):
    master_wb_gen.append([cell.value for cell in row])

daily_wb_brk=daily_wb['BreakDown Data'] #Generation sheet to be copied
master_wb_brk=master_wb['Breakdowns'] #to be copied into this sheet of master wb

for row in daily_wb_brk.iter_rows(min_row=2):
    master_wb_brk.append([cell.value for cell in row])

master_wb.save(dict_fname_change[latest_date])

##################################################################################################################################################################################################
# Data is automatically downloaded now data analysis part begins:

